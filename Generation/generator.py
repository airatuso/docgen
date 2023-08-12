from reportlab.lib.pagesizes import landscape, A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib.utils import ImageReader
import io
import textwrap
import openpyxl

# Загружаем данные из Excel файла
wb = openpyxl.load_workbook('Generation/data.xlsx')
sheet = wb.active

# Пути к шрифтам
font_fio_path = 'fonts/Gilroy-Bold.ttf'
font_program_path = 'fonts/Gilroy-Semibold.ttf'
font_number_path = 'fonts/Gilroy-Medium.ttf'

def generate_certificate(draw, fio_text, program_text, number_text):
    # Загружаем изображение-шаблон
    template_path = 'templates/background.png'
    template = Image.open(template_path)

    # Создаем объекты шрифтов
    font_fio = ImageFont.truetype(font_fio_path, size=150)
    font_program = ImageFont.truetype(font_program_path, size=100)
    font_number = ImageFont.truetype(font_number_path, size=62)

    # Цвета для каждого текста
    fio_color = (0, 0, 0)  # Черный цвет
    program_color = (32, 71, 152)  # Синий цвет
    number_color = (0, 0, 0)  # Черный цвет
    
    # Задаем координаты для каждого поля
    fio_position = (200, 1040)  # Замените на нужные координаты
    program_position = (200, 1550)  # Замените на нужные координаты
    number_position = (3210, 2267)  # Замените на нужные координаты
    
    # Создаем буфер для изображения
    img_buf = io.BytesIO()

    # Подставляем тексты и рисуем их на изображении
    draw_img = ImageDraw.Draw(template)
    
    # Разбиваем строку ФИО на фамилию и имя/отчество
    last_name, rest_of_name = fio_text.split(' ', 1)

    # Разбиваем строку программы обучения на несколько строк
    program_wrapped = textwrap.wrap(program_text, width=30)  

    # Рисуем тексты на изображении
    draw_img.text(fio_position, last_name, font=font_fio, fill=fio_color)
    draw_img.text((fio_position[0], fio_position[1] + 160), rest_of_name, font=font_fio, fill=fio_color)
    
    y_position = program_position[1]
    for line in program_wrapped:
        bbox = draw_img.textbbox(program_position, line, font=font_program)  # Получаем габаритный прямоугольник
        draw_img.text((program_position[0], y_position), line, font=font_program, fill=program_color)
        y_position += bbox[3] - bbox[1] + 30  # Добавляем небольшой отступ между строками
    
    draw_img.text(number_position, number_text, font=font_number, fill=number_color)

    # Сохраняем изображение в буфер
    template.save(img_buf, format='PNG')

    # Рисуем изображение на PDF
    img_buf.seek(0)
    draw.drawImage(ImageReader(img_buf), 0, 0, width=11.69*inch, height=8.27*inch)  # Размер A4 в дюймах

# Определите общее количество строк в таблице Excel
total_rows = sheet.max_row

# Создаем PDF документ формата А4 (горизонтальный)
pdf_output_path = 'certificates.pdf'
c = canvas.Canvas(pdf_output_path, pagesize=landscape(A4))

# Генерируем сертификаты для каждой строки данных
for row in range(2, total_rows + 1):  # Обходим все строки данных, начиная с второй строки
    fio_text = sheet.cell(row=row, column=1).value
    program_text = sheet.cell(row=row, column=2).value
    number_text = str(sheet.cell(row=row, column=3).value)
    generate_certificate(c, fio_text, program_text, number_text)
    c.showPage()

# Сохраняем PDF документ
c.save()
